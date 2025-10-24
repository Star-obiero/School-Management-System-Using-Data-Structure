import os
from openpyxl import load_workbook, Workbook


class LibrarySystem:
    def __init__(self, excel_file=r"D:\projects\DSassignment\School management\sampledata\library.xlsx"):
        self.excel_file = excel_file
        self.borrow_stack = []  # Stack for borrowed books (LIFO)
        self.books = {}         # book_id → {"title": ..., "status": ...}
        self.load_from_excel()

    # ---------------- EXCEL HANDLING ----------------
    def load_from_excel(self):
        """Load book data from Excel workbook."""
        if not os.path.exists(self.excel_file):
            print(f"⚠️ File '{self.excel_file}' not found. Creating a new one.")
            wb = Workbook()
            ws = wb.active
            ws.append(["book_id", "title", "status"])
            wb.save(self.excel_file)
            return

        wb = load_workbook(self.excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            book_id, title, status = row
            if book_id:
                self.books[book_id] = {"title": title, "status": status}

        wb.close()
        print(f"✅ Loaded {len(self.books)} books from {self.excel_file}.")

    def save_to_excel(self):
        """Save current book data back to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.append(["book_id", "title", "status"])

        for book_id, data in self.books.items():
            ws.append([book_id, data["title"], data["status"]])

        wb.save(self.excel_file)
        wb.close()
        print(f"💾 Library data saved to {self.excel_file}.")

    # ---------------- LIBRARY OPERATIONS ----------------
    def borrow_book(self, student_id, book_id):
        """Borrow a book and push it onto the stack."""
        if book_id not in self.books:
            print(f"⚠️ Book {book_id} not found.")
            return

        if self.books[book_id]["status"] == "Available":
            self.books[book_id]["status"] = f"Borrowed by {student_id}"
            self.borrow_stack.append(book_id)
            print(f"📚 {student_id} borrowed {book_id} - {self.books[book_id]['title']}.")
            self.save_to_excel()
        else:
            print(f"⚠️ Book {book_id} is already borrowed ({self.books[book_id]['status']}).")

    def return_book(self):
        """Return the last borrowed book (LIFO)."""
        if not self.borrow_stack:
            print("⚠️ No books to return — stack is empty.")
            return

        last_book = self.borrow_stack.pop()
        self.books[last_book]["status"] = "Available"
        print(f"✅ {last_book} - {self.books[last_book]['title']} returned successfully.")
        self.save_to_excel()

    def display_records(self):
        """Display all library books."""
        print("\n--- Library Books ---")
        for book_id, data in self.books.items():
            print(f"{book_id}: {data['title']} → {data['status']}")




if __name__ == "__main__":
    library = LibrarySystem(r"D:\projects\School management\sampledata\library.xlsx")

    library.display_records()

    # Borrow a book
    library.borrow_book("102", "BK-005")

    # Borrow another
    library.borrow_book("103", "BK-003")

    # Return the last borrowed book
    library.return_book()

    # Final view
    library.display_records()



